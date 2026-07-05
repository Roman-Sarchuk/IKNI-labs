class Alternative:
    def __init__(self, name: str, q1: int, q2: int):
        self.name = name
        self.q1 = q1
        self.q2 = q2
        self.pareto = None
        self.slater = None


def compare_pareto(a_main: Alternative, a_compare: Alternative):
    return ((a_main.q1 >= a_compare.q1 and a_main.q2 > a_compare.q2) or
            (a_main.q1 > a_compare.q1 and a_main.q2 >= a_compare.q2))


def compare_slater(a_main: Alternative, a_compare: Alternative):
    return a_main.q1 > a_compare.q1 and a_main.q2 > a_compare.q2


def get_answer_to_continue(answer_text, refusal_text):
    while True:
        print(f"\n{answer_text}? (Y/N)")
        answer = input("> ").lower().strip()

        if answer == "y":
            return True
        elif answer == "n":
            print(f"\n{refusal_text}!")
            return False

        print("\nЙойки, не коректний від, спробуйте ще раз!")


if __name__ == '__main__':
    # get information
    a_list = []
    while True:
        try:
            nums_string = input("Введіть числа через пробіл, наприклад: 05 04 33 99 34 19 15 і так далі\n> ")
            if not nums_string:
                print("Ви нічого не ввели!\n")
                continue
            nums_list = nums_string.split()
            a_list = [Alternative(name=f"A{i + 1}", q1=int(num[0]), q2=int(num[1]))
                      for i, num in enumerate(nums_list)]
            break
        except IndexError:
            print("Не коректний ввід!\n")

    # calc pareto & slater
    a_count = len(a_list)
    for i in range(a_count):
        for j in range(a_count):
            if i == j:
                continue

            if not a_list[j].pareto and compare_pareto(a_list[i], a_list[j]):
                a_list[j].pareto = f"A{i + 1}"

            if not a_list[j].slater and compare_slater(a_list[i], a_list[j]):
                a_list[j].slater = f"A{i + 1}"

    # show result in console & find pareto_res, slater_res
    pareto_res = {}     # *_res = {(x, y): [A3, A7, ..], ..}
    slater_res = {}     # *_res = {point: [alt.name, ..], ..}

    print("\nРезультати:")
    print("Alternative | Q1 | Q2 | Pareto | Slater")
    for alt in a_list:
        print(f"{alt.name:11} | {alt.q1:2} | {alt.q2:2} | {alt.pareto or '-':6} | {alt.slater or '-':6}")

        # Determination of optimal alternatives according to pareto and slater
        if not alt.pareto:
            pareto_res.setdefault((alt.q1, alt.q2), []).append(alt.name)
        if not alt.slater:
            slater_res.setdefault((alt.q1, alt.q2), []).append(alt.name)

    print(f"Pareto: " + ", ".join("=".join(alts_name) for alts_name in pareto_res.values()))
    print(f"Slater: " + ", ".join("=".join(alts_name) for alts_name in slater_res.values()))

    # transfer info into table
    if get_answer_to_continue("Зберегти дані у таблицю", "Ні, то ні. Буде без таблички"):
        import pandas as pd
        from openpyxl import load_workbook
        from openpyxl.styles import Border, Side, Alignment

        FILE_PATH = "multi_criteria_results.xlsx"

        try:
            data = {'/': ['Q1', 'Q2', 'Pareto', 'Slater']}
            for alt in a_list:
                # Example -> A6: [1, 9, A4, -]
                data[alt.name] = [alt.q1, alt.q2, alt.pareto or '-', alt.slater or '-']

            df = pd.DataFrame(data)
            print("\nЗбереження результатів у таблицю...")
            df.to_excel(FILE_PATH, index=False)
            print(f"Файл {FILE_PATH} збережено!")

            # Styling tabel
            print("\nСтилізація таблиці...")
            wb = load_workbook(FILE_PATH)
            ws = wb.active

            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            wb.save(FILE_PATH)
            print(f"Файл {FILE_PATH} збережено з форматуванням!")
        except PermissionError:
            print(f"ERROR: Файл {FILE_PATH} не доступний!\n\tМожливо даний файл зараз відкритий!")

    # visualize graphs
    if get_answer_to_continue("Намалювати графіки", "Ні, то ні. Буде без графіка"):
        import matplotlib.pyplot as plt
        import numpy as np


        def setup_subplot(ax, connected_points: tuple, color='r-', title=''):
            # Dictionary for grouping alternatives by coordinates
            point_coords = {}
            for alt in a_list:
                coords = (alt.q1, alt.q2)
                if coords not in point_coords:
                    point_coords[coords] = []
                point_coords[coords].append(alt)

            # Marking points while avoiding overlapping labels
            for coords, alts in point_coords.items():
                x, y = coords
                ax.scatter(x, y)

                # If there are several alternatives for this coordinate
                if len(alts) > 1:
                    # We place signatures in a circle
                    radius = 0.3
                    angles = np.linspace(0, 2 * np.pi, len(alts), endpoint=False)

                    for i, (alt, angle) in enumerate(zip(alts, angles)):
                        dx = radius * np.cos(angle)
                        dy = radius * np.sin(angle)
                        ax.text(x + dx, y + dy, alt.name, fontsize=9, ha='center', va='center')
                else:
                    # For a single point
                    ax.text(x + 0.2, y + 0.2, alts[0].name, fontsize=9)

            # Connecting the specified points with a line if more than one
            if connected_points and len(connected_points) > 1:
                # Sort by Q1 (increase), if equal - by Q2 (decrease)
                sorted_groups = sorted(connected_points, key=lambda p: (p[0], -p[1]))
                x_vals, y_vals = zip(*sorted_groups)  # Unpacking the coordinates

                # Draw a line
                ax.plot(x_vals, y_vals, color, linewidth=2, zorder=2)
            # Special marking for single points in connected_alts
            elif connected_points and len(connected_points) == 1:
                # Highlight the point with a larger marker and different colorn
                ax.scatter(connected_points[0][0], connected_points[0][1], s=150, facecolors='none',
                           edgecolors=color[0], linewidth=1)

            ax.set_title(title)
            ax.grid()


        # Create subplots
        fig, axs = plt.subplots(1, 2, figsize=(15, 7))

        # Graph 1
        setup_subplot(axs[0], tuple(pareto_res.keys()), 'r-', 'Границя Парето')

        # Graph 2
        setup_subplot(axs[1], tuple(slater_res.keys()), 'b-', 'Границя Слейтера')

        # Setting up subplots
        plt.tight_layout()
        plt.show()

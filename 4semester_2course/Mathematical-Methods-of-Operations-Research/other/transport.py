import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

# Вхідні дані для транспортної задачі
supply = [20, 30, 38, 42]
demand = [40, 30, 48, 12]
cost_matrix = [
    [6, 2, 4.8, 3],
    [8, 4, 5, 8],
    [5.5, 2, 3, 7],
    [1.8, 6, 7, 4]
]

# Приведення задачі до збалансованої
total_supply = sum(supply)
total_demand = sum(demand)

if total_supply > total_demand:
    demand.append(total_supply - total_demand)
    for row in cost_matrix:
        row.append(0)
elif total_demand > total_supply:
    supply.append(total_demand - total_supply)
    cost_matrix.append([0]*len(demand))

# Метод північно-західного кута
def northwest_corner(supply, demand, cost_matrix):
    s = supply.copy()
    d = demand.copy()
    allocation = np.zeros_like(cost_matrix)
    i = j = 0
    while i < len(s) and j < len(d):
        qty = min(s[i], d[j])
        allocation[i][j] = qty
        s[i] -= qty
        d[j] -= qty
        if s[i] == 0: i += 1
        if d[j] == 0: j += 1
    return allocation

# Метод мінімальних елементів
def least_cost(supply, demand, cost_matrix):
    s = supply.copy()
    d = demand.copy()
    allocation = np.zeros_like(cost_matrix)
    costs = [(i, j, cost_matrix[i][j]) for i in range(len(s)) for j in range(len(d))]
    costs.sort(key=lambda x: x[2])
    for i, j, _ in costs:
        if s[i] > 0 and d[j] > 0:
            qty = min(s[i], d[j])
            allocation[i][j] = qty
            s[i] -= qty
            d[j] -= qty
    return allocation

# Метод Фойгеля (наближення)
def vogels_approximation(supply, demand, cost_matrix):
    s = supply.copy()
    d = demand.copy()
    allocation = np.zeros_like(cost_matrix)
    cost = np.array(cost_matrix)

    while np.any(s) and np.any(d):
        penalties = []
        for i in range(len(s)):
            if s[i] > 0:
                row = [cost[i][j] for j in range(len(d)) if d[j] > 0]
                if len(row) >= 2:
                    sorted_row = sorted(row)
                    penalty = sorted_row[1] - sorted_row[0]
                elif len(row) == 1:
                    penalty = row[0]
                else:
                    penalty = -1
                penalties.append((penalty, 'row', i))
        for j in range(len(d)):
            if d[j] > 0:
                col = [cost[i][j] for i in range(len(s)) if s[i] > 0]
                if len(col) >= 2:
                    sorted_col = sorted(col)
                    penalty = sorted_col[1] - sorted_col[0]
                elif len(col) == 1:
                    penalty = col[0]
                else:
                    penalty = -1
                penalties.append((penalty, 'col', j))

        penalties = [p for p in penalties if p[0] >= 0]
        if not penalties:
            break

        penalties.sort(reverse=True)
        _, typ, idx = penalties[0]

        if typ == 'row':
            i = idx
            available = [(cost[i][j], j) for j in range(len(d)) if d[j] > 0]
            _, j = min(available)
        else:
            j = idx
            available = [(cost[i][j], i) for i in range(len(s)) if s[i] > 0]
            _, i = min(available)

        qty = min(s[i], d[j])
        allocation[i][j] = qty
        s[i] -= qty
        d[j] -= qty

    return allocation

# Розрахунок результатів
nw_alloc = northwest_corner(supply, demand, cost_matrix)
lc_alloc = least_cost(supply, demand, cost_matrix)
va_alloc = vogels_approximation(supply, demand, cost_matrix)

# Створення Excel-файлу
wb = Workbook()
ws1 = wb.active
ws1.title = "Northwest Corner"

df_nw = pd.DataFrame(nw_alloc, columns=[f'D{j+1}' for j in range(len(demand))], index=[f'S{i+1}' for i in range(len(supply))])
for r in dataframe_to_rows(df_nw, index=True, header=True):
    ws1.append(r)

ws2 = wb.create_sheet("Least Cost")
df_lc = pd.DataFrame(lc_alloc, columns=[f'D{j+1}' for j in range(len(demand))], index=[f'S{i+1}' for i in range(len(supply))])
for r in dataframe_to_rows(df_lc, index=True, header=True):
    ws2.append(r)

ws3 = wb.create_sheet("Vogels Approx.")
df_va = pd.DataFrame(va_alloc, columns=[f'D{j+1}' for j in range(len(demand))], index=[f'S{i+1}' for i in range(len(supply))])
for r in dataframe_to_rows(df_va, index=True, header=True):
    ws3.append(r)

# Збереження
file_path = "transportation_problem.xlsx"
wb.save(file_path)

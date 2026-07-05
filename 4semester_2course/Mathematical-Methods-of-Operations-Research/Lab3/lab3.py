import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from tabulate import tabulate
import random
import os


class InputDataHandler:
    FILE_NAME = "input_data.xlsx"

    def __init__(self):
        self._table = None

    def __str__(self):
        return tabulate(self._table, tablefmt="fancy_grid") if self._table else "data isn't read"

    def read_excel(self):
        print(f"Read data from the file '{self.FILE_NAME}'.....\n")
        try:
            df = pd.read_excel(self.FILE_NAME, header=None)

            self._verify_data(df)

            self._table = df.values.tolist()
        except FileNotFoundError as e:
            print(f"Failed to read the file:\n\t\tI can't find '{self.FILE_NAME}'.")
            print(f"Thus, I'll create such a file and fill '{self.FILE_NAME}' with random numbers in a 3x20 table.\n")
            self.init_file(2, 90)
            self.read_excel()
        except Exception as e:
            print(f"Failed to read the file:\n\t\t{e}")
            self._table = None

    @staticmethod
    def _verify_data(df):
        # Check if all rows have equal lengths
        row_lengths = df.apply(lambda row: len(row.dropna()), axis=1).tolist()
        if len(set(row_lengths)) != 1:
            raise ValueError("Rows have different lengths!")

        # Check if all elements are integers
        if not df.map(lambda x: isinstance(x, int)).all().all():
            raise ValueError("The file has no integer values!")

    def init_file(self, min_value, max_value):
        wb = Workbook()
        ws = wb.active

        # Заповнення таблиці випадковими числами
        for row in range(1, 4):  # 3 рядки
            for col in range(1, 21):  # 20 стовпців
                # Генерація випадкового числа в заданому діапазоні
                random_number = random.randint(min_value, max_value)

                # Запис числа в комірку
                ws.cell(row=row, column=col, value=random_number)

        # Збереження файлу
        wb.save(self.FILE_NAME)

    def get(self):
        return self._table


class OutputDataHandler:
    FILE_NAME = "output_table.xlsx"

    def __init__(self):
        self.data = []
        self.init_data()

    def init_data(self):
        self.data = [
            ["Дані", "Кількість контейнерів", "", "", "", "Обчислювальна складність", "", "", ""],
            ["", "Без впорядкування", "", "", "", "Без порядкування", "", "", ""],
            ["", "NFA", "FFA", "WFA", "BFA", "NFA", "FFA", "WFA", "BFA"],
            ["1 рядок"],
            ["2 рядок"],
            ["3 рядок"],
            ["1+2+3 рядок"],
            ["Дані", "З впорядкуванням", "", "", "", "З порядкуванням", "", "", ""],
            ["", "NFA", "FFA", "WFA", "BFA", "NFA", "FFA", "WFA", "BFA"],
            ["1 рядок"],
            ["2 рядок"],
            ["3 рядок"],
            ["1+2+3 рядок"]
        ]

    def add_row(self, sort: bool, row_num: int, row: list):
        if row_num <= 0 or row_num > 4:
            raise ValueError("row_num is from 1 to 4")

        index = 8 if sort else 2
        index += row_num

        if len(self.data[index]) > 1:
            raise ValueError(f"row in index #{index} is filled already")

        self.data[index] += row

    def save(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Аркуш1"

        # Writing data to a table
        for row in self.data:
            ws.append(row)

        # Merging cells
        ws.merge_cells("A1:A3")
        ws.merge_cells("B1:E1")
        ws.merge_cells("F1:I1")
        ws.merge_cells("B2:E2")
        ws.merge_cells("F2:I2")
        ws.merge_cells("A8:A9")
        ws.merge_cells("B8:E8")
        ws.merge_cells("F8:I8")

        # Formatting cells
        bold_font = Font(bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style="thin"),
                             right=Side(style="thin"),
                             top=Side(style="thin"),
                             bottom=Side(style="thin"))
        gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

        for row in ws.iter_rows(min_row=1, max_row=len(self.data), min_col=1, max_col=9):
            for cell in row:
                cell.alignment = center_align
                cell.border = thin_border
                if cell.row in [1, 2, 3, 8, 9]:
                    cell.font = bold_font
                    cell.fill = gray_fill

        try:
            wb.save(self.FILE_NAME)
        except PermissionError:
            print(f"Failed to open the file '{self.FILE_NAME}'\n\t\tClose this file and rerun.")
        else:
            print(f"Result successfully saved in the file '{self.FILE_NAME}'.")


def quicksort_with_counter(arr):
    # Лічильник циклів
    cycles = [0]

    def quick_sort_recursive(arr, low, high):
        cycles[0] += 1

        if low < high:
            # Знаходимо індекс розділення
            partition_index = partition(arr, low, high)

            # Рекурсивно сортуємо ліву та праву частини
            quick_sort_recursive(arr, low, partition_index - 1)
            quick_sort_recursive(arr, partition_index + 1, high)

    def partition(arr, low, high):
        # Обираємо останній елемент як опорний
        pivot = arr[high]

        # Індекс меншого елементу
        i = low - 1

        for j in range(low, high):
            cycles[0] += 1
            # Якщо поточний елемент менший або рівний опорному
            if arr[j] >= pivot:
                # Інкрементуємо індекс меншого елементу
                i += 1
                arr[i], arr[j] = arr[j], arr[i]

        # Розміщуємо опорний елемент у правильну позицію
        arr[i + 1], arr[high] = arr[high], arr[i + 1]
        return i + 1

    # Створюємо копію масиву для сортування
    sorted_arr = arr.copy()

    # Викликаємо рекурсивну функцію сортування
    quick_sort_recursive(sorted_arr, 0, len(sorted_arr) - 1)

    return sorted_arr, cycles[0]


def efficient_sort_with_index(arr):
    """
    Ефективна функція сортування для невеликих масивів.

    :param arr: Вхідний масив чисел
    :return: Кортеж (sorted_enumerated_array, cycle_count)
    """
    # Перевірка вхідних даних
    if not arr:
        return [], 0

    # Лічильник циклів
    cycles = 0

    # Створення масиву кортежів (значення, оригінальний_індекс)
    enumerated_arr = [{"value": value, "index": index} for index, value in enumerate(arr)]

    # Insertion sort для невеликих масивів
    for i in range(1, len(enumerated_arr)):
        key = enumerated_arr[i]
        j = i - 1

        # Порівняння та переміщення елементів
        while j >= 0 and enumerated_arr[j]["value"] > key["value"]:
            enumerated_arr[j + 1] = enumerated_arr[j]
            j -= 1
            cycles += 1

        enumerated_arr[j + 1] = key
        cycles += 1

    return enumerated_arr, cycles


class NPTable:
    FILE_NAME = "process_table.xlsx"

    def __init__(self):
        self.table_data = []
        self.n = 20
        self.is_able_to_open_file = True
        self.delete_file()

    def init(self, n):
        self.n = n
        self.table_data = [["#"] + list(range(1, self.n + 1))]

    def add_new_row(self):
        self.table_data.append([len(self.table_data)] + [""] * self.n)

    def set(self, row, col, value):
        """
        :param row: from 1 to self.n
        :param col: from 1 to len(self.table_data)
        :param value: number
        """
        self.table_data[row][col] = value

    def save_new_sheet(self, sheet_name):
        if not self.is_able_to_open_file:
            return

        try:
            wb = load_workbook(self.FILE_NAME)
        except FileNotFoundError:
            wb = Workbook()
            # Delete default sheet
            wb.remove(wb.active)

        ws = wb.create_sheet(title=sheet_name)

        for row in self.table_data:
            ws.append(row)

        try:
            wb.save(self.FILE_NAME)
        except PermissionError:
            self.is_able_to_open_file = False
            print(f"Failed to open the file '{self.FILE_NAME}'\n\t\tClose this file and rerun.")

    def delete_file(self):
        if not self.is_able_to_open_file:
            return

        try:
            # Перевірка, чи файл існує
            if not os.path.exists(self.FILE_NAME):
                print(f"Файл {self.FILE_NAME} не існує.")
                return False

            # Видалення файлу
            os.remove(self.FILE_NAME)

            return True
        except PermissionError:
            self.is_able_to_open_file = False
            print(f"Failed to open the file '{self.FILE_NAME}'\n\t\tClose this file and rerun.")
        except Exception as e:
            print(f"Failed to delete the file: {e}")
            return False


class NPAlgorithm:
    def __init__(self, table, container_capacity, saving=False):
        self.table = table
        self.capacity = container_capacity
        self.algorithms = [self.nfa, self.ffa, self.wfa, self.bfa]
        # Variables to saving
        self.res_table = NPTable()
        self.sort = False
        self.row_num = 0
        self.process_count = 0
        self.saving = True

    def is_save_successful(self):
        if not self.saving:
            return None

        return self.res_table.is_able_to_open_file

    def get_row_result(self, sort: bool, row_index: int):
        """
        :return: list = [nfa, ffa, wfa, bfa, counter_nfa, counter_ffa, counter_wfa, counter_bfa]
        """
        if row_index < 0 or row_index > 2:
            raise ValueError("row_num is from 0 to 2")

        self.row_num = row_index + 1
        return self._process(self.table[row_index], sort)

    def get_table_result(self, sort: bool):
        """
        :return: list = [nfa, ffa, wfa, bfa, counter_nfa, counter_ffa, counter_wfa, counter_bfa]
        """
        self.row_num = "1-3"
        return self._process(sum(self.table, []), sort)

    def _process(self, weights: list, sort: bool):
        """
        :return: list = [nfa, ffa, wfa, bfa, counter_nfa, counter_ffa, counter_wfa, counter_bfa]
        """
        sort_counter = 0
        if sort:
            self.sort = sort
            weights, sort_counter = quicksort_with_counter(weights)

        result = []
        for i in range(len(self.algorithms)):
            self.process_count += 1
            containers, counter = self.algorithms[i](weights)
            result.insert(i, containers)
            result.append(counter + sort_counter)

        return result

    def nfa(self, weights):
        """
        :return: containers, counter
        """
        self.res_table.init(len(weights))
        self.res_table.add_new_row()
        container_count = 1
        comparisons = 0
        fullness = 0

        for i in range(len(weights)):
            comparisons += 1

            total = fullness + weights[i]
            if total <= 100:
                fullness = total
            else:
                self.res_table.add_new_row()
                container_count += 1
                fullness = weights[i]
            self.res_table.set(container_count, i + 1, weights[i])

        if self.saving:
            self.res_table.save_new_sheet(f"{self.row_num}.{'s.' if self.sort else ''}NFA.{self.process_count}")
        return container_count, comparisons

    def ffa(self, weights):
        """
        :return: containers, counter
        """
        self.res_table.init(len(weights))
        self.res_table.add_new_row()
        containers = [0]
        comparisons = 0

        container_id = -1
        for i in range(len(weights)):
            comparisons += 1

            total = containers[-1] + weights[i]
            if total <= 100:
                containers[-1] = total
                container_id = len(containers)
            else:
                j = len(containers) - 2
                while j >= 0:
                    comparisons += 1

                    total = containers[j] + weights[i]
                    if total <= 100:
                        containers[j] = total
                        container_id = j + 1
                        break
                    j -= 1
                if j < 0:
                    self.res_table.add_new_row()
                    containers.append(weights[i])
                    container_id = len(containers)
            self.res_table.set(container_id, i + 1, weights[i])

        if self.saving:
            self.res_table.save_new_sheet(f"{self.row_num}.{'s.' if self.sort else ''}FFA.{self.process_count}")
        return len(containers), comparisons

    def wfa(self, weights):
        """
        :return: containers, counter
        """
        self.res_table.init(len(weights))
        self.res_table.add_new_row()
        containers = [0]
        comparisons = 0

        container_id = -1
        for i in range(len(weights)):
            comparisons += 1

            total = containers[-1] + weights[i]
            if total <= 100:
                containers[-1] = total
                container_id = len(containers)
            else:
                min_index = 0
                for j in range(1, len(containers)):
                    comparisons += 1
                    if containers[j] < containers[min_index]:
                        min_index = j

                total = containers[min_index] + weights[i]
                if total <= 100:
                    containers[min_index] = total
                    container_id = min_index + 1
                else:
                    self.res_table.add_new_row()
                    containers.append(weights[i])
                    container_id = len(containers)
            self.res_table.set(container_id, i + 1, weights[i])

        if self.saving:
            self.res_table.save_new_sheet(f"{self.row_num}.{'s.' if self.sort else ''}WFA.{self.process_count}")
        return len(containers), comparisons

    def bfa(self, weights):
        """
        :return: containers, counter
        """
        self.res_table.init(len(weights))
        self.res_table.add_new_row()
        containers = [0]
        comparisons = 0

        container_id = -1
        for i in range(len(weights)):
            comparisons += 1

            total = containers[-1] + weights[i]
            if total <= 100:
                containers[-1] = total
                container_id = len(containers)
            else:
                containers_sorted, comparisons_sort = efficient_sort_with_index(containers)
                comparisons += comparisons_sort

                j = len(containers_sorted) - 1
                while j >= 0:
                    comparisons += 1

                    total = containers_sorted[j]["value"] + weights[i]
                    if total <= 100:
                        containers[containers_sorted[j]["index"]] = total
                        container_id = containers_sorted[j]["index"] + 1
                        break
                    j -= 1
                if j < 0:
                    self.res_table.add_new_row()
                    containers.append(weights[i])
                    container_id = len(containers)
            self.res_table.set(container_id, i + 1, weights[i])

        if self.saving:
            self.res_table.save_new_sheet(f"{self.row_num}.{'s.' if self.sort else ''}BFA.{self.process_count}")
        return len(containers), comparisons

    def min_containers_estimate(self, weights):
        """
        :return: min_containers_estimate (int)
        """
        return -(-sum(weights) // self.capacity)


if __name__ == '__main__':
    CAPACITY = 100

    # get data
    input_data_handler = InputDataHandler()
    input_data_handler.read_excel()
    data = input_data_handler.get()

    if data:
        # show received data
        print("Received data successfully:\n", input_data_handler, sep="")

        print("\nProcessing.....\n")
        algorithm_handler = NPAlgorithm(data, CAPACITY)
        output_data_handler = OutputDataHandler()

        output_data_handler.add_row(False, 1, algorithm_handler.get_row_result(False, 0))
        output_data_handler.add_row(False, 2, algorithm_handler.get_row_result(False, 1))
        output_data_handler.add_row(False, 3, algorithm_handler.get_row_result(False, 2))
        output_data_handler.add_row(False, 4, algorithm_handler.get_table_result(False))

        output_data_handler.add_row(True, 1, algorithm_handler.get_row_result(True, 0))
        output_data_handler.add_row(True, 2, algorithm_handler.get_row_result(True, 1))
        output_data_handler.add_row(True, 3, algorithm_handler.get_row_result(True, 2))
        output_data_handler.add_row(True, 4, algorithm_handler.get_table_result(True))

        if algorithm_handler.is_save_successful() is not False:
            print("Result saving.....")
            output_data_handler.save()
